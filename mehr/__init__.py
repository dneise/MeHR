from datetime import datetime, timedelta
from datetime import time as dt_time
import requests
import openpyxl
import os
import time
import iso8601
from dateutil.parser import parse
from commandr import command, Run

from .iso_to_country import iso_to_country
from .config import load_config


@command('repeat')
def repeat():
    config = load_config()
    period = timedelta(hours=24).total_seconds()
    next_execution = 0
    while True:
        current_time = time.time()
        if current_time >= next_execution:
            mews_report = get_started_reservations_yesterday(config)
            rows = mews_report_to_report_rows(mews_report)
            write_excel_output_file(
                rows,
                outfolder=config.OutFolder
            )
            next_execution = current_time + period


@command('main')
def date(when='22.06.2018'):
    config = load_config()
    mews_report = get_started_reservations_yesterday(
        config,
        start_utc=None if when is None else parse(when).date()
    )
    rows = mews_report_to_report_rows(mews_report)
    write_excel_output_file(
        rows,
        outfolder=config.OutFolder
    )


def get_started_reservations_yesterday(
    config,
    time_filter='Start',
    start_utc=None,
    end_utc=None,
    states=['Started', ],
    extent={
        "Reservations": False,
        "ReservationGroups": False,
        "Customers": True,
        "Spaces": True,
    },
):

    if start_utc is None:
        start_utc = today_midnight() - timedelta(days=1)
    if end_utc is None:
        end_utc = start_utc + timedelta(days=100)
    return reservations_getAll(
        config,
        time_filter=time_filter,
        start_utc=start_utc,
        end_utc=end_utc,
        states=states,
        extent=extent
    )


def today_midnight():
    return datetime.combine(datetime.utcnow().date(), dt_time())


def reservations_getAll(
    config,
    time_filter=None,
    start_utc=datetime.utcnow(),
    end_utc=None,
    states=None,
    extent=None,
):
    '''
    time_filter - string, default Colliding
    start_utc - string, default: Now
    end_utc - string, default: Now + 1 day
    states - array of string, default ['Confirmed', 'Started', 'Processed']
    extent - string, default Reservations, Groups and Customers
    currency - string
    '''

    if end_utc is None:
        end_utc = start_utc + timedelta(days=1)

    return requests.post(
        '{PlatformAddress}/api/connector/v1/{Resource}/{Operation}'.format(
            PlatformAddress=config.PlatformAddress,
            Resource='reservations',
            Operation='getAll',
        ),
        json={
            "ClientToken": config.ClientToken,
            "AccessToken": config.AccessToken,
            "LanguageCode": None,
            "CultureCode": None,
            "TimeFilter": time_filter,
            "StartUtc": start_utc.isoformat(),
            "EndUtc": end_utc.isoformat()
        }
    ).json()


def mews_report_to_report_rows(mews_report):
    rows = []
    customers = {
        customer['Id']: customer
        for customer in mews_report['Customers']
    }
    if mews_report['Spaces'] is not None:
        spaces = {
            space['Id']: space
            for space in mews_report['Spaces']
        }
    else:
        spaces = {}

    for reservation in mews_report['Reservations']:
        row = {}
        customer = customers[reservation['CustomerId']]
        space = spaces.get(reservation['AssignedSpaceId'], None)

        if customer['BirthDateUtc'] is not None:
            birth_date = to_date(customer['BirthDateUtc'])
            row.update({
                'Geboren Tag': str(birth_date.day),
                'Monat': str(birth_date.month),
                'Jahr': str(birth_date.year),
            })
        else:
            row.update({
                'Geboren Tag': '',
                'Monat': '',
                'Jahr': '',
            })

        row.update({
            'Meldeschein Nr.': str(len(rows) + 1),
            'Familienname': customer['LastName'],
            'Vornamen': customer['FirstName'],
            'Staatsangehörigkeit ISO': str(customer['NationalityCode']),
            'Staatsangehörigkeit': iso_to_country.get((customer['NationalityCode']), ''),
            'Ankunft': to_date(reservation['StartUtc']).strftime('%d.%m.%Y'),
            'Abreise': to_date(reservation['EndUtc']).strftime('%d.%m.%Y'),
            'Zimmernummer': 'unknown' if space is None else space['Number'],
            'Ausweisnummer': ausweisnummer_from_customer(customer)
        })
        rows.append(row)
    return rows


def write_excel_output_file(rows, outfolder):
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "HoKo"
    for col_index, col_name in enumerate(HOKO_EXCEL_REPORT_COLUMN_NAMES):
        sh.cell(row=1, column=col_index+1, value=col_name)

    for row_id, row in enumerate(rows):
        row_number = row_id + 2
        for col_index, col_name in enumerate(HOKO_EXCEL_REPORT_COLUMN_NAMES):

            sh.cell(
                row=row_number,
                column=col_index+1,
                value=row[col_name]
            )

    if not os.path.isdir(outfolder):
        print('''\
The outfolder you specified in your config file does not exist.
OutFolder: {outfolder}
I am going to create that folder now.''')
        os.makedirs(outfolder)
    wb.save(os.path.join(outfolder, 'test.xls'))


HOKO_EXCEL_REPORT_COLUMN_NAMES = [
    'Meldeschein Nr.',
    'Zimmernummer',
    'Familienname',
    'Vornamen',
    'Geboren Tag',
    'Monat',
    'Jahr',
    'Staatsangehörigkeit',
    'Staatsangehörigkeit ISO',
    'Ausweisnummer',
    'Ankunft',
    'Abreise'
]


def zimmernummer_from_reservation(reservation):
    raise NotImplementedError


def ausweisnummer_from_customer(customer):
    ''' customers can identify with different documents
    we have a certain order of how much we like different documents
    so when a customer identifies with multiple documents we only give the
    number of the document we like best.
    '''
    DOCUMENT_NAME_ORDER = [
        'Passport',
        'IdentityCard',
        'Visa',
        'DriversLicense'
    ]
    for document_name in DOCUMENT_NAME_ORDER:
        if customer[document_name] is not None:
            number = customer[document_name]['Number'].strip()
            # with the if below we just make sure that customers having
            # by accident or human error multiple documents, but some
            # without a number, are not identified with a passport with an
            # empty number as long as they have a drivers license with a valid
            # number.
            # Not sure if this is needed, depends on business logic on the
            # mews side, which I do not know.
            if number:
                return number
    return ''


def to_date(s):
    return iso8601.parse_date(s)


def entry():
    Run(main='main')


if __name__ == '__main__':
    entry()
