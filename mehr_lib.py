from datetime import datetime, timedelta
from datetime import time as dt_time
import requests
import openpyxl
import os
import os.path
import iso8601

from hoko_iso_to_country import iso_to_country


def last_midnight():
    return datetime.combine(datetime.utcnow().date(), dt_time())


def reservations_getAll(
    config,
    time_filter='Start',
    start_utc=None,
    end_utc=None,
    states=['Started', ],
    extent={
        "Customers": True,
        "Reservations": True,
        "Spaces": True,
    },
):
    '''
    time_filter - string, default Colliding
    start_utc - string, default: Now
    end_utc - string, default: Now + 1 day
    states - array of string, default ['Confirmed', 'Started', 'Processed']
    extent - string, default Reservations, Groups and Customers
    currency - string
    '''
    if start_utc is None:
        start_utc = last_midnight() - timedelta(days=1)
    if end_utc is None:
        end_utc = start_utc + timedelta(days=1)

    response = requests.post(
        '{PlatformAddress}/api/connector/v1/{Resource}/{Operation}'.format(
            PlatformAddress=config.PlatformAddress,
            Resource='reservations',
            Operation='getAll',
        ),
        json={
            "ClientToken": config.ClientToken,
            "AccessToken": config.AccessToken,
            "LanguageCode": None,
            "CultureCode": None,
            "TimeFilter": time_filter,
            "StartUtc": start_utc.isoformat(),
            "EndUtc": end_utc.isoformat(),
            "Extent": extent,
        }
    )
    return response.json(), start_utc


def mews_report_to_report_rows(mews_report):
    rows = []
    customers = {
        customer['Id']: customer
        for customer in mews_report['Customers']
    }
    if mews_report['Spaces'] is not None:
        spaces = {
            space['Id']: space
            for space in mews_report['Spaces']
        }
    else:
        spaces = {}

    for reservation in mews_report['Reservations']:
        row = {}
        customer = customers[reservation['CustomerId']]
        space = spaces.get(reservation['AssignedSpaceId'], None)

        if customer['BirthDateUtc'] is not None:
            birth_date = iso8601.parse_date(customer['BirthDateUtc'])
            row.update({
                'Geboren Tag': str(birth_date.day),
                'Monat': str(birth_date.month),
                'Jahr': str(birth_date.year),
            })
        else:
            row.update({
                'Geboren Tag': '',
                'Monat': '',
                'Jahr': '',
            })

        row.update({
            'Meldeschein Nr.': str(len(rows) + 1),
            'Familienname': customer['LastName'],
            'Vornamen': customer['FirstName'],
            'Staatsangehörigkeit ISO': str(customer['NationalityCode']),
            'Staatsangehörigkeit': iso_to_country.get((customer['NationalityCode']), ''),
            'Ankunft': iso8601.parse_date(
                reservation['StartUtc']
            ).strftime('%d.%m.%Y'),
            'Abreise': iso8601.parse_date(
                reservation['EndUtc']
            ).strftime('%d.%m.%Y'),
            'Zimmernummer': 'unknown' if space is None else space['Number'],
            'Ausweisnummer': ausweisnummer_from_customer(customer)
        })
        rows.append(row)
    return rows


def write_excel_output_file(rows, outpath):
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "HoKo"
    from openpyxl.cell.cell import get_column_letter
    for i in range(len(HOKO_EXCEL_REPORT_COLUMN_NAMES)):
        sh.column_dimensions[get_column_letter(i+1)].width = 15

    my_fill = openpyxl.styles.fills.PatternFill(
        patternType='solid',
        fgColor=openpyxl.styles.colors.Color(rgb='0099ccff')
    )
    for col_index, col_name in enumerate(HOKO_EXCEL_REPORT_COLUMN_NAMES):
        sh.cell(row=1, column=col_index+1, value=col_name)
        sh.cell(
            row=1,
            column=col_index+1
        ).fill = my_fill

    for row_id, row in enumerate(rows):
        row_number = row_id + 2
        for col_index, col_name in enumerate(HOKO_EXCEL_REPORT_COLUMN_NAMES):

            sh.cell(
                row=row_number,
                column=col_index+1,
                value=row[col_name]
            )

    outfolder = os.path.dirname(outpath)
    if not os.path.isdir(outfolder):
        os.makedirs(outfolder)
    wb.save(outpath)


HOKO_EXCEL_REPORT_COLUMN_NAMES = [
    'Meldeschein Nr.',
    'Zimmernummer',
    'Familienname',
    'Vornamen',
    'Geboren Tag',
    'Monat',
    'Jahr',
    'Staatsangehörigkeit',
    'Staatsangehörigkeit ISO',
    'Ausweisnummer',
    'Ankunft',
    'Abreise'
]


def ausweisnummer_from_customer(customer):
    ''' customers can identify with different documents
    we have a certain order of how much we like different documents
    so when a customer identifies with multiple documents we only give the
    number of the document we like best.
    '''
    DOCUMENT_NAME_ORDER = [
        'Passport',
        'IdentityCard',
        'Visa',
        'DriversLicense'
    ]
    for document_name in DOCUMENT_NAME_ORDER:
        if customer[document_name] is not None:
            number = customer[document_name]['Number'].strip()
            # with the if below we just make sure that customers having
            # by accident or human error multiple documents, but some
            # without a number, are not identified with a passport with an
            # empty number as long as they have a drivers license with a valid
            # number.
            # Not sure if this is needed, depends on business logic on the
            # mews side, which I do not know.
            if number:
                return number
    return ''
