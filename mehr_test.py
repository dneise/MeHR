from datetime import datetime, timedelta
from datetime import time as dt_time
import requests
import openpyxl
import os
import os.path
from types import SimpleNamespace
import iso8601
from dateutil.parser import parse

import json

here = os.path.dirname(os.path.realpath(__file__))

config_template = {
    'PlatformAddress': "https://demo.mews.li",
    'ClientToken': "E0D439EE522F44368DC78E1BFB03710C-D24FB11DBE31D4621C4817E028D9E1D",
    'Hotels': [
        {
            'Name': "Test Hotel",
            'AccessToken': "C66EF7B239D24632943D115EDE9CB810-EA00F8FD8294692C940F6B5A8F9453D",
            'OutFolder': None,
            'FileName': 'testhotel_%Y%m%d.xls'
        },
    ]
}


config_not_found_message = '''\
There is no "config.json" file in {here}
I am going to make one for you now.
Please, adjust the file to your personal needs and start me again.
'''


iso_to_country = {
 'AD': 'Andorra',
 'AE': 'Vereinigte arabische Emirate',
 'AF': 'Afghanistan',
 'AG': 'Antigua und Barbuda',
 'AI': 'Anguilla',
 'AL': 'Albanien',
 'AM': 'Armenien',
 'AN': 'Niederländische Antillen',
 'AO': 'Angola',
 'AQ': 'Antarktis',
 'AR': 'Argentinien',
 'AS': 'Amerikanisch Samoa',
 'AT': 'Oesterreich',
 'AU': 'Australien',
 'AW': 'Aruba',
 'AZ': 'Aserbaidschan',
 'BA': 'Bosnien und Herzegowina',
 'BB': 'Barbados',
 'BD': 'Bangladesh',
 'BE': 'Belgien',
 'BF': 'Burkina Faso',
 'BG': 'Bulgarien',
 'BH': 'Bahrain',
 'BI': 'Burundi',
 'BJ': 'Benin',
 'BM': 'Bermuda (GB)',
 'BN': 'Brunei',
 'BO': 'Bolivien',
 'BR': 'Brasilien',
 'BS': 'Bahamas',
 'BT': 'Bhutan',
 'BV': 'Bouvet-Insel',
 'BW': 'Botswana',
 'BY': 'Belarus (Weissrussland)',
 'BZ': 'Belize',
 'CA': 'Kanada',
 'CC': 'Cocos-Inseln',
 'CF': 'Zentralafrikanische Republik',
 'CG': 'Kongo',
 'CH': 'Schweiz',
 'CI': 'Elfenbeinküste',
 'CK': 'Cook-Inseln',
 'CL': 'Chile',
 'CM': 'Kamerun',
 'CN': 'China (Volksrepublik)',
 'CO': 'Kolumbien',
 'CR': 'Costa Rica',
 'CU': 'Kuba',
 'CV': 'Kapverden',
 'CX': 'Christmas-Insel (AUS)',
 'CY': 'Zypern',
 'CZ': 'Tschechische Republik',
 'DE': 'Deutschland',
 'DJ': 'Djibouti',
 'DK': 'Dänemark',
 'DM': 'Dominica',
 'DO': 'Dominikanische Republik',
 'DZ': 'Algerien',
 'EC': 'Ecuador',
 'EE': 'Estland',
 'EG': 'Aegypten',
 'EH': 'Westsahara',
 'ER': 'Eritrea',
 'ES': 'Spanien',
 'ET': 'Aethiopien',
 'FI': 'Finnland',
 'FJ': 'Fidschi (Inseln)',
 'FK': 'Falkland-Inseln',
 'FM': 'Mikronesien',
 'FR': 'Frankreich',
 'GA': 'Gabun',
 'GB': 'Grossbritannien, Vereinigtes Königreich',
 'GD': 'Grenada',
 'GE': 'Georgien',
 'GF': 'Französisch-Guyana',
 'GH': 'Ghana',
 'GI': 'Gibraltar',
 'GL': 'Grönland',
 'GM': 'Gambia',
 'GN': 'Guinea',
 'GP': 'Guadeloupe',
 'GQ': 'Aequatorial-Guinea',
 'GR': 'Griechenland',
 'GT': 'Guatemala',
 'GU': 'Guam',
 'GW': 'Guinea-Bissau',
 'GY': 'Guyana (Republik)',
 'HK': 'Hongkong',
 'HM': 'Heard-, MacDonald-Inseln',
 'HN': 'Honduras',
 'HR': 'Kroatien',
 'HT': 'Haiti',
 'HU': 'Ungarn',
 'ID': 'Indonesien',
 'IE': 'Irland',
 'IL': 'Israel',
 'IN': 'Indien',
 'IO': 'Britische Territorien im indischen Ozean',
 'IQ': 'Irak',
 'IR': 'Iran',
 'IS': 'Island',
 'ISO': 'Land',
 'IT': 'Italien',
 'JM': 'Jamaika',
 'JO': 'Jordanien',
 'JP': 'Japan',
 'KE': 'Kenia',
 'KG': 'Kirgisistan',
 'KH': 'Kambodscha',
 'KI': 'Kiribati',
 'KM': 'Komoren',
 'KN': 'St. Kitts und Nevis',
 'KO': 'Kosovo',
 'KP': 'Korea (Nord)',
 'KR': 'Korea (Süd)',
 'KW': 'Kuwait',
 'KY': 'Kaiman-Inseln GB',
 'KZ': 'Kasachstan',
 'LA': 'Laos',
 'LB': 'Libanon',
 'LC': 'St. Lucia',
 'LI': 'Liechtenstein',
 'LK': 'Sri Lanka',
 'LR': 'Liberia',
 'LS': 'Lesotho',
 'LT': 'Litauen',
 'LU': 'Luxemburg',
 'LV': 'Lettland',
 'LY': 'Libyen',
 'MA': 'Marokko',
 'MC': 'Monaco',
 'MD': 'Moldau',
 'ME': 'Montenegro',
 'MG': 'Madagaskar',
 'MH': 'Marshall-Inseln',
 'MK': 'Mazedonien',
 'ML': 'Mali',
 'MM': 'Myanmar',
 'MN': 'Mongolei',
 'MO': 'Macao',
 'MQ': 'Martinique',
 'MR': 'Mauretanien',
 'MS': 'Montserrat',
 'MT': 'Malta',
 'MU': 'Mauritius',
 'MV': 'Malediven',
 'MW': 'Malawi',
 'MX': 'Mexiko',
 'MY': 'Malaysia',
 'MZ': 'Mosambik',
 'NA': 'Namibia',
 'NC': 'Neukaledonien',
 'NE': 'Niger',
 'NF': 'Norfolk-Insel',
 'NG': 'Nigeria',
 'NI': 'Nicaragua',
 'NL': 'Niederlande',
 'NO': 'Norwegen',
 'NP': 'Nepal',
 'NR': 'Nauru',
 'NU': 'Niue',
 'NZ': 'Neuseeland',
 'OM': 'Oman',
 'PA': 'Panama',
 'PE': 'Peru',
 'PF': 'Französisch-Polynesien',
 'PG': 'Papua-Neuguinea',
 'PH': 'Philippinen',
 'PK': 'Pakistan',
 'PL': 'Polen',
 'PM': 'St-Pierre und Miquelon',
 'PN': 'Pitcairn-Inseln',
 'PR': 'Puerto Rico',
 'PS': 'Palästina',
 'PT': 'Portugal',
 'PW': 'Palau',
 'PY': 'Paraguay',
 'QA': 'Katar',
 'RE': 'Reunion',
 'RO': 'Rumänien',
 'RS': 'Serbien',
 'RU': 'Russland',
 'RW': 'Rwanda',
 'SA': 'Saudi Arabien',
 'SB': 'Salomon Inseln',
 'SC': 'Seychellen',
 'SD': 'Sudan',
 'SE': 'Schweden',
 'SG': 'Singapur',
 'SH': 'St. Helena',
 'SI': 'Slowenien',
 'SJ': 'Svalbard und Jan Mayen',
 'SK': 'Slowakische Republik',
 'SL': 'Sierra Leone',
 'SM': 'San Marino',
 'SN': 'Senegal',
 'SO': 'Somalia',
 'SR': 'Suriname',
 'ST': 'São Tome und Principe',
 'SV': 'El Salvador',
 'SY': 'Syrien',
 'SZ': 'Swasiland',
 'TC': 'Turks-, Caicos-Inseln',
 'TD': 'Tschad',
 'TF': 'Südliches Eismeer (F)',
 'TG': 'Togo',
 'TH': 'Thailand',
 'TJ': 'Tadschikistan',
 'TK': 'Tokelau',
 'TM': 'Turkmenistan',
 'TN': 'Tunesien',
 'TO': 'Tonga',
 'TP': 'Timor-Leste',
 'TR': 'Türkei',
 'TT': 'Trinidad und Tobago',
 'TV': 'Tuvalu',
 'TW': 'Taiwan (China)',
 'TZ': 'Tansania',
 'UA': 'Ukraine',
 'UB': 'Officional of special organisation',
 'UG': 'Uganda',
 'UK': 'Unmik (Kosovo)',
 'UM': 'Treuhandinseln (USA)',
 'UN': 'Vereinte Nationen',
 'US': 'Vereinigte Staaten von Amerika (USA)',
 'UY': 'Uruguay',
 'UZ': 'Usbekistan',
 'VA': 'Vatikanstadt',
 'VC': 'St. Vincent und die Grenadinen',
 'VE': 'Venezuela',
 'VG': 'Jungfern-Inseln (UK)',
 'VI': 'Jungfern-Inseln (USA)',
 'VN': 'Vietnam',
 'VU': 'Vanuatu (Republik)',
 'WF': 'Wallis und Futuna',
 'WS': 'Samoa (West)',
 'X': 'unbekannt/inconnu/sconosciuto',
 'XP': 'INTERPOL',
 'XX': 'Staatenlos',
 'YE': 'Jemen',
 'YT': 'Mayotte (Insel)',
 'YU': 'Bundesrepublik Jugoslawien',
 'ZA': 'Südafrika',
 'ZM': 'Sambia',
 'ZR': 'Demokratische Republik Kongo',
 'ZW': 'Zimbabwe',
 'ZZ': 'Nation unbekannt (Personenfahndung)'
}


def load_config(path_to_config=here):
    config_path = os.path.join(path_to_config, 'config.json')
    if not os.path.isfile(config_path):
        print(config_not_found_message.format(here=path_to_config))
        config_template['Hotels'][0]['OutFolder'] = path_to_config
        json.dump(
            config_template,
            open(config_path, 'w'),
            indent=4,
        )
    config = json.load(open(config_path))
    configs = flatten_config(config)
    return configs


def flatten_config(cfg):
    configs = []
    for hotel in cfg['Hotels']:
        hotel['PlatformAddress'] = cfg['PlatformAddress']
        hotel['ClientToken'] = cfg['ClientToken']
    configs.append(hotel)
    return configs


def date(when='22.06.2018'):
    configs = load_config()
    for config in configs:
        config = SimpleNamespace(**config)
        mews_report = get_started_reservations_yesterday(
            config,
            start_utc=None if when is None else parse(when).date()
        )
        rows = mews_report_to_report_rows(mews_report)

        now = datetime.now()
        outpath = os.path.join(config.OutFolder, now.strftime(config.FileName))
        write_excel_output_file(
            rows,
            outpath=outpath
        )


def get_started_reservations_yesterday(
    config,
    time_filter='Start',
    start_utc=None,
    end_utc=None,
    states=['Started', ],
    extent={
        "Customers": True,
        "Reservations": True,
        "Spaces": True,
    },
):

    if start_utc is None:
        start_utc = today_midnight() - timedelta(days=1)
    if end_utc is None:
        end_utc = start_utc + timedelta(days=1)
    return reservations_getAll(
        config,
        time_filter=time_filter,
        start_utc=start_utc,
        end_utc=end_utc,
        states=states,
        extent=extent
    )


def today_midnight():
    return datetime.combine(datetime.utcnow().date(), dt_time())


def reservations_getAll(
    config,
    time_filter=None,
    start_utc=datetime.utcnow(),
    end_utc=None,
    states=None,
    extent=None,
):
    '''
    time_filter - string, default Colliding
    start_utc - string, default: Now
    end_utc - string, default: Now + 1 day
    states - array of string, default ['Confirmed', 'Started', 'Processed']
    extent - string, default Reservations, Groups and Customers
    currency - string
    '''

    if end_utc is None:
        end_utc = start_utc + timedelta(days=1)

    response = requests.post(
        '{PlatformAddress}/api/connector/v1/{Resource}/{Operation}'.format(
            PlatformAddress=config.PlatformAddress,
            Resource='reservations',
            Operation='getAll',
        ),
        json={
            "ClientToken": config.ClientToken,
            "AccessToken": config.AccessToken,
            "LanguageCode": None,
            "CultureCode": None,
            "TimeFilter": time_filter,
            "StartUtc": start_utc.isoformat(),
            "EndUtc": end_utc.isoformat(),
            "Extent": extent,
        }
    )
    return response.json()


def mews_report_to_report_rows(mews_report):
    rows = []
    customers = {
        customer['Id']: customer
        for customer in mews_report['Customers']
    }
    if mews_report['Spaces'] is not None:
        spaces = {
            space['Id']: space
            for space in mews_report['Spaces']
        }
    else:
        spaces = {}

    for reservation in mews_report['Reservations']:
        row = {}
        customer = customers[reservation['CustomerId']]
        space = spaces.get(reservation['AssignedSpaceId'], None)

        if customer['BirthDateUtc'] is not None:
            birth_date = to_date(customer['BirthDateUtc'])
            row.update({
                'Geboren Tag': str(birth_date.day),
                'Monat': str(birth_date.month),
                'Jahr': str(birth_date.year),
            })
        else:
            row.update({
                'Geboren Tag': '',
                'Monat': '',
                'Jahr': '',
            })

        row.update({
            'Meldeschein Nr.': str(len(rows) + 1),
            'Familienname': customer['LastName'],
            'Vornamen': customer['FirstName'],
            'Staatsangehörigkeit ISO': str(customer['NationalityCode']),
            'Staatsangehörigkeit': iso_to_country.get((customer['NationalityCode']), ''),
            'Ankunft': to_date(reservation['StartUtc']).strftime('%d.%m.%Y'),
            'Abreise': to_date(reservation['EndUtc']).strftime('%d.%m.%Y'),
            'Zimmernummer': 'unknown' if space is None else space['Number'],
            'Ausweisnummer': ausweisnummer_from_customer(customer)
        })
        rows.append(row)
    return rows


def write_excel_output_file(rows, outpath):
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "HoKo"
    from openpyxl.cell.cell import get_column_letter
    for i in range(len(HOKO_EXCEL_REPORT_COLUMN_NAMES)):
        sh.column_dimensions[get_column_letter(i+1)].width = 15

    my_fill = openpyxl.styles.fills.PatternFill(
        patternType='solid',
        fgColor=openpyxl.styles.colors.Color(rgb='0099ccff')
    )
    for col_index, col_name in enumerate(HOKO_EXCEL_REPORT_COLUMN_NAMES):
        sh.cell(row=1, column=col_index+1, value=col_name)
        sh.cell(
            row=1,
            column=col_index+1
        ).fill = my_fill

    for row_id, row in enumerate(rows):
        row_number = row_id + 2
        for col_index, col_name in enumerate(HOKO_EXCEL_REPORT_COLUMN_NAMES):

            sh.cell(
                row=row_number,
                column=col_index+1,
                value=row[col_name]
            )

    outfolder = os.path.dirname(outpath)
    if not os.path.isdir(outfolder):
        os.makedirs(outfolder)
    wb.save(outpath)


HOKO_EXCEL_REPORT_COLUMN_NAMES = [
    'Meldeschein Nr.',
    'Zimmernummer',
    'Familienname',
    'Vornamen',
    'Geboren Tag',
    'Monat',
    'Jahr',
    'Staatsangehörigkeit',
    'Staatsangehörigkeit ISO',
    'Ausweisnummer',
    'Ankunft',
    'Abreise'
]


def zimmernummer_from_reservation(reservation):
    raise NotImplementedError


def ausweisnummer_from_customer(customer):
    ''' customers can identify with different documents
    we have a certain order of how much we like different documents
    so when a customer identifies with multiple documents we only give the
    number of the document we like best.
    '''
    DOCUMENT_NAME_ORDER = [
        'Passport',
        'IdentityCard',
        'Visa',
        'DriversLicense'
    ]
    for document_name in DOCUMENT_NAME_ORDER:
        if customer[document_name] is not None:
            number = customer[document_name]['Number'].strip()
            # with the if below we just make sure that customers having
            # by accident or human error multiple documents, but some
            # without a number, are not identified with a passport with an
            # empty number as long as they have a drivers license with a valid
            # number.
            # Not sure if this is needed, depends on business logic on the
            # mews side, which I do not know.
            if number:
                return number
    return ''


def to_date(s):
    return iso8601.parse_date(s)


if __name__ == '__main__':
    date()
